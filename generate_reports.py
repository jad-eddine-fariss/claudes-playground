#!/usr/bin/env python3
"""
Generate one standalone formulation report sheet per Fo_ID from
FORMULATIONS_DATABASE_V2.xlsm (read-only) into a new workbook
IS_all_formulation_reports.xlsx.

Read-only on the master: only openpyxl.load_workbook(..., data_only=True) is
ever called on it; a brand new Workbook() is written for the output.
"""
import re
import math
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MASTER = "FORMULATIONS_DATABASE_V2.xlsm"
OUTPUT = "IS_all_formulation_reports.xlsx"
GENERATED_DATE = datetime.date(2026, 8, 26)

GREEN_HEADER = "01B24B"
GREEN_LABEL = "007431"
LIGHT_FILL = "EAF7EE"
GRAY_TEXT = "666666"
BORDER_GRAY = "CCCCCC"
FONT_NAME = "Montserrat"

ACTIVES = ['N', 'P2O5', 'K2O', 'S', 'Ca', 'Mg', 'Zn', 'B', 'Mn', 'Cu', 'Mo',
           'Se', 'Fe', 'Si', 'Humic acid', 'Fulvic acid', 'Seaweed extract']

NEEDS_INPUT_KEYWORDS = ['to be filled']
CONFIRM_KEYWORDS = ['please confirm']
DEDUCED_KEYWORDS = ['target', 'deduc', 'dilut', 'confirm', 'from zno', 'renormalis']

NO_LONG_DASH = str.maketrans({'—': '-', '–': '-'})

N_COLS = 9  # A..I
COL_WIDTHS = [22, 26, 14, 12, 10, 12, 10, 12, 40]

thin = Side(style='thin', color=BORDER_GRAY)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)


def clean(s):
    if s is None:
        return None
    return str(s).translate(NO_LONG_DASH)


def load_master():
    print("Loading master workbook (read-only, data_only=True)...")
    wb = openpyxl.load_workbook(MASTER, data_only=True, keep_vba=False, read_only=False)
    return wb


def sheet_to_dicts(ws):
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        rowvals = [ws.cell(row=r, column=i + 1).value for i in range(len(headers))]
        if all(v is None for v in rowvals):
            continue
        d = {h: v for h, v in zip(headers, rowvals)}
        d['_row'] = r
        rows.append(d)
    return rows


def extract():
    wb = load_master()
    summary_rows = sheet_to_dicts(wb['Summary'])
    formulation_rows = sheet_to_dicts(wb['Formulations'])

    fo_order = []
    summary_by_fo = {}
    for row in summary_rows:
        fo = row['Fo_ID']
        if fo is None:
            continue
        fo_order.append(fo)
        summary_by_fo[fo] = row

    formulations_by_fo = {}
    for row in formulation_rows:
        fo = row['Fo_ID']
        if fo is None:
            continue
        formulations_by_fo.setdefault(fo, []).append(row)

    for fo, rows in formulations_by_fo.items():
        rows.sort(key=lambda r: (r.get('Step') if r.get('Step') is not None else 0, r['_row']))

    wb.close()
    return fo_order, summary_by_fo, formulations_by_fo


# ---------- classification ----------

def find_active_in_text(text):
    if not text:
        return []
    found = []
    for a in sorted(ACTIVES, key=len, reverse=True):
        if a in text:
            found.append(a)
    return found


def classify_fo(fo, summary, frows):
    flag = clean(summary.get('Flag'))
    needs_input_items = []
    row_missing_dose_ids = set()
    guarantee_pending_detail = {}

    for r in frows:
        if r.get('Amount') is None:
            row_missing_dose_ids.add(r['_row'])
            ing = r.get('Ingredient', '?')
            reason = f' ({flag})' if flag else ''
            needs_input_items.append(f"{ing}: dose/amount not recorded{reason}")

    if flag and any(k in flag.lower() for k in NEEDS_INPUT_KEYWORDS):
        actives_named = find_active_in_text(flag)
        raw = clean(summary.get('Raw / self-active')) or ''
        ing_names = [seg.split(' (')[0].strip() for seg in raw.split(';') if seg.strip()]
        ing_label = ' / '.join(ing_names) if ing_names else 'ingredient'
        for a in actives_named:
            guarantee_pending_detail[a] = ing_label
        active_label = '/'.join(actives_named) if actives_named else 'active'
        needs_input_items.append(
            f"{ing_label}: {active_label} guarantee not filled in Guarantees sheet ({flag})")

    if flag and any(k in flag.lower() for k in CONFIRM_KEYWORDS):
        needs_input_items.append(f'Confirm assumption in source Flag: "{flag}"')

    if needs_input_items:
        completeness = 'needs_input'
        completeness_note = "Needs input: " + "; ".join(needs_input_items)
    elif flag and any(k in flag.lower() for k in DEDUCED_KEYWORDS):
        fields = set()
        for r in frows:
            if clean(r.get('Flag / interpretation')) == flag and r.get('Role') != 'Carrier':
                for a in ACTIVES:
                    v = r.get(f'{a} in blend')
                    if v:
                        fields.add(a)
        if not fields:
            for a in ACTIVES:
                v = summary.get(f'{a} in final blend')
                if v:
                    fields.add(a)
        fields_str = '/'.join(sorted(fields)) if fields else 'recipe ratio (see flag note)'
        completeness = 'deduced'
        completeness_note = f"Deduced: {fields_str} (see Flag / interpretation note above)"
    else:
        completeness = 'complete'
        completeness_note = "Complete"

    return {
        'completeness': completeness,
        'completeness_note': completeness_note,
        'needs_input_items': needs_input_items,
        'row_missing_dose_ids': row_missing_dose_ids,
        'guarantee_pending_detail': guarantee_pending_detail,
        'flag': flag,
    }


def parse_self_active(text):
    if not text:
        return []
    items = []
    for seg in text.split(';'):
        seg = seg.strip()
        if not seg:
            continue
        m = re.match(r'^(.*)\s(-?\d+(?:\.\d+)?)\s*%$', seg)
        if m:
            items.append((m.group(1).strip(), m.group(2)))
        else:
            items.append((seg, None))
    return items


# ---------- sub-formulation resolution ----------
#
# A formulation's recipe can use another formulation as an ingredient
# (Role == 'Sub-formulation', e.g. Fo51 <- Fo50 at Pct_of_blend of Fo51's
# blend). The master's own Summary!"X in final blend" figures already fully
# and correctly resolve the tracked mineral/biostimulant ACTIVES through
# these chains (verified against an independent bottom-up recomputation
# across all 197 formulations x 17 actives: zero mismatches, including
# multi-level chains). The one field the master does NOT propagate through
# sub-formulation references is the free-text "Raw / self-active" list
# (non-tracked raw materials like coatings/biopolymers/acids) - a parent
# formulation's own Summary row simply omits raw/self-active items that
# only exist inside a referenced sub-formulation. This resolver fixes that
# by recursively pulling in the sub-formulation's own (already-resolved)
# raw/self-active composition, scaled by that row's Pct_of_blend - the
# same mass-fraction logic the master already applies to the actives.

_self_active_cache = {}


def resolve_self_active_fracs(fo, summary_by_fo, formulations_by_fo, fo_ids, stack=()):
    """Fo's own effective raw/self-active composition as {name: fraction of
    this fo's total blend}, including whatever it inherits (recursively)
    from any Sub-formulation ingredient rows. Memoized; read-only derivation
    from master data, no invented values."""
    if fo in _self_active_cache:
        return _self_active_cache[fo]
    s = summary_by_fo.get(fo, {})
    result = {}
    for name, pct_str in parse_self_active(clean(s.get('Raw / self-active'))):
        if pct_str is not None:
            result[name] = result.get(name, 0.0) + float(pct_str) / 100.0
    if fo not in stack:
        for row in formulations_by_fo.get(fo, []):
            if str(row.get('Role') or '').strip().lower() == 'sub-formulation':
                sub_fo = row.get('Ingredient')
                row_pct = row.get('Pct_of_blend') or 0
                if sub_fo in fo_ids:
                    sub_fracs = resolve_self_active_fracs(
                        sub_fo, summary_by_fo, formulations_by_fo, fo_ids, stack + (fo,))
                    for name, frac in sub_fracs.items():
                        result[name] = result.get(name, 0.0) + row_pct * frac
    _self_active_cache[fo] = result
    return result


def compute_self_active_render(fo, summary, frows, summary_by_fo, formulations_by_fo, fo_ids):
    """List of (name, display_pct_or_None, note_or_None) for the Raw /
    self-active block, folding in sub-formulation contributions. Returns
    None when there is nothing to add, signalling the caller to use the
    formulation's own text unchanged (covers the 144/197 formulations with
    no sub-formulation row or whose sub-formulation carries no raw/self-
    active material)."""
    own_items = parse_self_active(clean(summary.get('Raw / self-active')))

    sub_rows = [r for r in frows if str(r.get('Role') or '').strip().lower() == 'sub-formulation']
    if not sub_rows:
        return None

    extra = {}
    extra_via = {}
    for r in sub_rows:
        sub_fo = r.get('Ingredient')
        row_pct = r.get('Pct_of_blend') or 0
        if sub_fo not in fo_ids:
            continue
        sub_fracs = resolve_self_active_fracs(sub_fo, summary_by_fo, formulations_by_fo, fo_ids)
        for name, frac in sub_fracs.items():
            contributed = row_pct * frac
            if contributed <= 0:
                continue
            extra[name] = extra.get(name, 0.0) + contributed
            extra_via.setdefault(name, set()).add(sub_fo)

    if not extra:
        return None

    result = []
    own_names = set()
    for name, pct_str in own_items:
        own_names.add(name)
        if name in extra:
            via = ', '.join(sorted(extra_via[name]))
            if pct_str is None:
                result.append((name, "not recorded (direct)",
                                f"Direct amount not recorded; additionally contains content via "
                                f"sub-formulation {via} (scaled by its share of this blend)."))
            else:
                total = float(pct_str) / 100.0 + extra[name]
                result.append((name, f"{total * 100:.2f}%",
                                f"Direct {pct_str}% plus content via sub-formulation {via} "
                                f"(scaled by its share of this blend)."))
        else:
            result.append((name, (f"{pct_str}%" if pct_str is not None else None), None))
    for name, frac in extra.items():
        if name in own_names:
            continue
        via = ', '.join(sorted(extra_via[name]))
        result.append((name, f"{frac * 100:.2f}%",
                        f"Via sub-formulation {via} (scaled by its share of this blend); "
                        f"not a direct ingredient of {fo}."))
    return result


# ---------- styling helpers ----------

def f_title():
    return Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")


def f_section():
    return Font(name=FONT_NAME, size=12, bold=True, color=GREEN_LABEL)


def f_label():
    return Font(name=FONT_NAME, size=10, bold=True, color=GREEN_LABEL)


def f_value(bold=False, italic=False, color="000000"):
    return Font(name=FONT_NAME, size=10, bold=bold, italic=italic, color=color)


def f_table_header():
    return Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")


def f_footer():
    return Font(name=FONT_NAME, size=9, italic=True, color=GRAY_TEXT)


def set_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def estimate_height(text, chars_per_line=128, line_h=15, pad=6, min_h=18):
    if not text:
        return min_h
    lines = str(text).split('\n')
    total = sum(max(1, math.ceil(len(line) / chars_per_line)) for line in lines)
    return max(min_h, total * line_h + pad)


def merge_row(ws, row, c1, c2):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)


def write_title(ws, row, text):
    merge_row(ws, row, 1, N_COLS)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = f_title()
    cell.fill = PatternFill(fill_type='solid', fgColor=GREEN_HEADER)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 28
    for c in range(2, N_COLS + 1):
        ws.cell(row=row, column=c).fill = PatternFill(fill_type='solid', fgColor=GREEN_HEADER)


def write_label_value(ws, row, label, value, wrap=False):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = f_label()
    lc.alignment = Alignment(horizontal='left', vertical='top')
    merge_row(ws, row, 2, N_COLS)
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = f_value()
    if wrap:
        vc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = estimate_height(value, chars_per_line=150)
    else:
        vc.alignment = Alignment(horizontal='left', vertical='top')
    return row + 1


def write_section(ws, row, text):
    merge_row(ws, row, 1, N_COLS)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = f_section()
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=0)
    cell.border = Border(bottom=Side(style='medium', color=GREEN_HEADER))
    for c in range(2, N_COLS + 1):
        ws.cell(row=row, column=c).border = Border(bottom=Side(style='medium', color=GREEN_HEADER))
    ws.row_dimensions[row].height = 20
    return row + 1


def write_note_line(ws, row, text):
    merge_row(ws, row, 1, N_COLS)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = f_value(italic=True, color=GRAY_TEXT)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = estimate_height(text, chars_per_line=150)
    return row + 1


# ---------- report sheet ----------

def build_report_sheet(out_wb, fo, summary, frows, cls, summary_by_fo=None, formulations_by_fo=None, fo_ids=None):
    ws = out_wb.create_sheet(title=fo)
    set_col_widths(ws)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    row = 1
    write_title(ws, row, "FORMULATION RECORD")
    row += 1
    row += 1  # spacer

    dry_wet = summary.get('Dry_Wet') or 'Dry'
    date_val = summary.get('Date')
    date_str = date_val.strftime('%d %b %Y') if isinstance(date_val, datetime.datetime) else (clean(date_val) or 'not specified')

    row = write_label_value(ws, row, "Fo_ID", fo)
    row = write_label_value(ws, row, "Date", date_str)
    row = write_label_value(ws, row, "Site", clean(summary.get('Site')) or 'not specified')
    row = write_label_value(ws, row, "Made by", clean(summary.get('Made_by')) or 'not specified')
    row = write_label_value(ws, row, "Stakeholder BU", clean(summary.get('Stakeholder_BU')) or 'not specified')
    row = write_label_value(ws, row, "Process", clean(summary.get('Process')) or 'not specified')
    row = write_label_value(ws, row, "Dry / Wet", dry_wet)

    flag = cls['flag']
    if flag:
        row = write_label_value(ws, row, "Flag / interpretation", flag, wrap=True)

    comp_color = {"complete": "1E7B34", "deduced": "9C6500", "needs_input": "B00020"}[cls['completeness']]
    lc = ws.cell(row=row, column=1, value="Data completeness")
    lc.font = f_label()
    merge_row(ws, row, 2, N_COLS)
    vc = ws.cell(row=row, column=2, value=cls['completeness_note'])
    vc.font = f_value(bold=True, color=comp_color)
    vc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = estimate_height(cls['completeness_note'], chars_per_line=150)
    row += 1

    row += 1  # spacer

    # ---- Raw materials (the recipe) ----
    row = write_section(ws, row, "RAW MATERIALS (RECIPE)")
    headers = ["Step", "Ingredient", "Role", "Amount", "Basis", "Lab_g (g)", "kg/T", "% of blend", "Note"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = f_table_header()
        cell.fill = PatternFill(fill_type='solid', fgColor=GREEN_LABEL)
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 24
    row += 1

    missing_ids = cls['row_missing_dose_ids']
    for r in frows:
        is_missing = r['_row'] in missing_ids
        ing_note = clean(r.get('Ingredient note'))
        note_parts = []
        if ing_note:
            note_parts.append(ing_note)
        if is_missing:
            note_parts.append(f"not recorded ({cls['flag']})" if cls['flag'] else "not recorded")
        note_text = "; ".join(note_parts)

        values = [
            r.get('Step'),
            clean(r.get('Ingredient')),
            clean(r.get('Role')),
            ("not recorded" if is_missing else r.get('Amount')),
            (clean(r.get('Basis')) or '-'),
            ("not recorded" if is_missing else r.get('Lab_g')),
            ("not recorded" if is_missing else r.get('kg_T')),
            ("not recorded" if (is_missing or r.get('Pct_of_blend') is None) else r.get('Pct_of_blend')),
            note_text,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = f_value()
            cell.border = BORDER_ALL
            cell.alignment = Alignment(horizontal='center' if c in (1, 4, 6, 7, 8) else 'left',
                                        vertical='center', wrap_text=(c in (2, 9)))
            if c == 8 and isinstance(v, (int, float)):
                cell.number_format = '0.000%'
            if c in (4, 6) and isinstance(v, (int, float)):
                cell.number_format = '0.####'
        row += 1

    row += 1  # spacer

    # ---- Final product - active content ----
    row = write_section(ws, row, "FINAL PRODUCT - ACTIVE CONTENT")
    basis_label = ("Guarantees in the finished product (dried granule basis - water excluded)"
                   if dry_wet == 'Dry' else
                   "Guarantees in the finished product (includes water)")
    row = write_note_line(ws, row, basis_label)

    guarantee_pending = cls['guarantee_pending_detail']
    non_zero_actives = [a for a in ACTIVES if summary.get(f'{a} in final blend')]
    combined = list(dict.fromkeys(non_zero_actives + list(guarantee_pending.keys())))

    if not combined:
        if cls['row_missing_dose_ids']:
            row = write_note_line(row=row, ws=ws,
                                   text="Active content not computable: ingredient dose(s) not recorded (see Needs input / Data completeness above).")
        else:
            row = write_note_line(row=row, ws=ws,
                                   text="No broken-down mineral/biostimulant actives. This formulation is 100% self-active/raw-material components (see below).")
    else:
        act_headers = ["Active", "% in final product", "Note"]
        for c, h in enumerate(act_headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = f_table_header()
            cell.fill = PatternFill(fill_type='solid', fgColor=GREEN_LABEL)
            cell.border = BORDER_ALL
        merge_row(ws, row, 3, N_COLS)
        ws.row_dimensions[row].height = 18
        row += 1
        for a in combined:
            label = f"{a} (%)"
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = f_value()
            lc.border = BORDER_ALL
            if a in guarantee_pending:
                vc = ws.cell(row=row, column=2, value="guarantee pending")
                vc.font = f_value(bold=True, color="B00020")
                val = summary.get(f'{a} in final blend') or 0
                note = (f"Partial value from recorded source only: {val:.2%}. "
                        f"{guarantee_pending[a]} guarantee not yet filled in Guarantees sheet - "
                        f"true total will be higher once filled.")
            else:
                val = summary.get(f'{a} in final blend') or 0
                vc = ws.cell(row=row, column=2, value=val)
                vc.number_format = '0.000%'
                vc.font = f_value()
                note = ""
            vc.border = BORDER_ALL
            vc.alignment = Alignment(horizontal='center')
            merge_row(ws, row, 3, N_COLS)
            nc = ws.cell(row=row, column=3, value=note)
            nc.font = f_value(italic=True, color=GRAY_TEXT)
            nc.border = BORDER_ALL
            nc.alignment = Alignment(horizontal='left', wrap_text=True)
            if note:
                ws.row_dimensions[row].height = estimate_height(note, chars_per_line=150)
            row += 1

    resolved_self_active = None
    if summary_by_fo is not None and formulations_by_fo is not None and fo_ids is not None:
        resolved_self_active = compute_self_active_render(
            fo, summary, frows, summary_by_fo, formulations_by_fo, fo_ids)

    if resolved_self_active is not None:
        self_active_items = resolved_self_active  # [(name, display_pct_or_None, note_or_None), ...]
    else:
        raw_active_text = clean(summary.get('Raw / self-active'))
        self_active_items = [(name, (f"{pct}%" if pct is not None else None), None)
                              for name, pct in parse_self_active(raw_active_text)]

    if self_active_items:
        lc = ws.cell(row=row, column=1, value="Raw / self-active")
        lc.font = f_label()
        merge_row(ws, row, 2, N_COLS)
        ws.cell(row=row, column=2)
        row += 1
        for name_cat, display_pct, note in self_active_items:
            if cls['row_missing_dose_ids'] and display_pct in ('0%', '0.0%', '0.00%'):
                display_pct = "not recorded"
            elif display_pct is None:
                display_pct = "not recorded"
            lc = ws.cell(row=row, column=1, value=f"  {name_cat}")
            lc.font = f_value()
            if note:
                merge_row(ws, row, 2, 3)
                vc = ws.cell(row=row, column=2, value=display_pct)
                vc.font = f_value()
                merge_row(ws, row, 4, N_COLS)
                nc = ws.cell(row=row, column=4, value=note)
                nc.font = f_value(italic=True, color=GRAY_TEXT)
                nc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.row_dimensions[row].height = estimate_height(note, chars_per_line=90)
            else:
                merge_row(ws, row, 2, N_COLS)
                vc = ws.cell(row=row, column=2, value=display_pct)
                vc.font = f_value()
            row += 1

    row += 1  # spacer

    # ---- Protocol ----
    row = write_section(ws, row, "PROTOCOL")
    protocol_text = clean(summary.get('Protocol')) or 'not recorded'
    merge_row(ws, row, 1, N_COLS)
    cell = ws.cell(row=row, column=1, value=protocol_text)
    cell.font = f_value()
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = estimate_height(protocol_text, chars_per_line=150)
    row += 1

    protocol_check = clean(summary.get('Protocol check')) or 'not recorded'
    row = write_label_value(ws, row, "Protocol check", protocol_check)

    row += 1  # spacer

    # ---- Footer ----
    basis_line = ("Dry: percentages on dried granule" if dry_wet == 'Dry'
                  else "Wet: percentages include water")
    merge_row(ws, row, 1, N_COLS)
    fcell = ws.cell(row=row, column=1,
                     value=f"Generated: {GENERATED_DATE.strftime('%d %b %Y')}    |    Basis: {basis_line}")
    fcell.font = f_footer()
    row += 1

    ws.print_area = f"A1:I{row}"
    return ws


# ---------- index & needs-input sheets ----------

def build_index_sheet(out_wb, fo_order, summary_by_fo, classifications, tally):
    ws = out_wb.create_sheet(title="Index", index=0)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([12, 14, 16, 16, 12, 10, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    merge_row(ws, row, 1, 7)
    c = ws.cell(row=row, column=1, value="FORMULATION REPORTS INDEX")
    c.font = f_title()
    c.fill = PatternFill(fill_type='solid', fgColor=GREEN_HEADER)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for cc in range(2, 8):
        ws.cell(row=row, column=cc).fill = PatternFill(fill_type='solid', fgColor=GREEN_HEADER)
    ws.row_dimensions[row].height = 28
    row += 2

    ws.cell(row=row, column=1, value=f"Generated: {GENERATED_DATE.strftime('%d %b %Y')}").font = f_footer()
    row += 1
    ws.cell(row=row, column=1,
            value=f"{len(fo_order)} formulation reports  |  Complete: {tally['complete']}  |  "
                  f"Deduced: {tally['deduced']}  |  Needs input: {tally['needs_input']}").font = f_value(bold=True, color=GREEN_LABEL)
    row += 1
    ws.cell(row=row, column=1,
            value="Source: FORMULATIONS_DATABASE_V2.xlsm (read-only). See the 'Needs Input' sheet for exact gaps.").font = f_footer()
    row += 2

    headers = ["Fo_ID", "Date", "Site", "Process", "Dry/Wet", "Lines", "Data completeness"]
    for cidx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=cidx, value=h)
        cell.font = f_table_header()
        cell.fill = PatternFill(fill_type='solid', fgColor=GREEN_LABEL)
        cell.border = BORDER_ALL
    row += 1
    header_row = row - 1

    comp_color = {"complete": "1E7B34", "deduced": "9C6500", "needs_input": "B00020"}
    for fo in fo_order:
        s = summary_by_fo[fo]
        cls = classifications[fo]
        date_val = s.get('Date')
        date_str = date_val.strftime('%d %b %Y') if isinstance(date_val, datetime.datetime) else (clean(date_val) or '')
        vals = [fo, date_str, clean(s.get('Site')) or '', clean(s.get('Process')) or '',
                clean(s.get('Dry_Wet')) or '', s.get('# ingredient lines'), cls['completeness_note']]
        for cidx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=cidx, value=v)
            cell.font = f_value(color=(comp_color[cls['completeness']] if cidx == 7 else "000000"),
                                 bold=(cidx == 7))
            cell.border = BORDER_ALL
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cidx == 7))
        # hyperlink Fo_ID to its sheet
        link_cell = ws.cell(row=row, column=1)
        link_cell.hyperlink = f"#'{fo}'!A1"
        link_cell.font = Font(name=FONT_NAME, size=10, color="0563C1", underline='single')
        row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.print_area = f"A1:G{row}"


def build_needs_input_sheet(out_wb, fo_order, classifications):
    ws = out_wb.create_sheet(title="Needs Input")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([12, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    merge_row(ws, row, 1, 2)
    c = ws.cell(row=row, column=1, value="NEEDS INPUT - ITEMS TO CLOSE")
    c.font = f_title()
    c.fill = PatternFill(fill_type='solid', fgColor=GREEN_HEADER)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.cell(row=row, column=2).fill = PatternFill(fill_type='solid', fgColor=GREEN_HEADER)
    ws.row_dimensions[row].height = 28
    row += 2

    needs = [fo for fo in fo_order if classifications[fo]['completeness'] == 'needs_input']
    ws.cell(row=row, column=1,
            value=f"{len(needs)} of {len(fo_order)} formulations have at least one unresolved gap.").font = f_value(bold=True, color=GREEN_LABEL)
    row += 2

    headers = ["Fo_ID", "Exact missing field(s)"]
    for cidx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=cidx, value=h)
        cell.font = f_table_header()
        cell.fill = PatternFill(fill_type='solid', fgColor=GREEN_LABEL)
        cell.border = BORDER_ALL
    row += 1

    for fo in needs:
        cls = classifications[fo]
        link_cell = ws.cell(row=row, column=1, value=fo)
        link_cell.hyperlink = f"#'{fo}'!A1"
        link_cell.font = Font(name=FONT_NAME, size=10, color="0563C1", underline='single', bold=True)
        link_cell.border = BORDER_ALL
        link_cell.alignment = Alignment(vertical='top')

        text = "\n".join(f"- {item}" for item in cls['needs_input_items'])
        vc = ws.cell(row=row, column=2, value=text)
        vc.font = f_value()
        vc.border = BORDER_ALL
        vc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = estimate_height(text, chars_per_line=95, line_h=14)
        row += 1

    ws.freeze_panes = "A5"


# ---------- main ----------

def main():
    fo_order, summary_by_fo, formulations_by_fo = extract()
    print(f"Total formulations: {len(fo_order)}")

    classifications = {}
    for fo in fo_order:
        classifications[fo] = classify_fo(fo, summary_by_fo[fo], formulations_by_fo.get(fo, []))

    tally = {'complete': 0, 'deduced': 0, 'needs_input': 0}
    for fo in fo_order:
        tally[classifications[fo]['completeness']] += 1
    print("Tally:", tally)

    out_wb = openpyxl.Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    fo_ids = set(summary_by_fo.keys())
    for i, fo in enumerate(fo_order, start=1):
        build_report_sheet(out_wb, fo, summary_by_fo[fo], formulations_by_fo.get(fo, []), classifications[fo],
                            summary_by_fo=summary_by_fo, formulations_by_fo=formulations_by_fo, fo_ids=fo_ids)
        if i % 20 == 0 or i == len(fo_order):
            print(f"  ... {i}/{len(fo_order)} report sheets built")

    build_index_sheet(out_wb, fo_order, summary_by_fo, classifications, tally)
    build_needs_input_sheet(out_wb, fo_order, classifications)

    out_wb.save(OUTPUT)
    print(f"Saved {OUTPUT}")

    with open('run_summary.txt', 'w') as f:
        f.write(f"Total formulations: {len(fo_order)}\n")
        f.write(f"Complete: {tally['complete']}\n")
        f.write(f"Deduced: {tally['deduced']}\n")
        f.write(f"Needs input: {tally['needs_input']}\n\n")
        f.write("Needs input detail:\n")
        for fo in fo_order:
            cls = classifications[fo]
            if cls['completeness'] == 'needs_input':
                f.write(f"{fo}:\n")
                for item in cls['needs_input_items']:
                    f.write(f"  - {item}\n")
        f.write("\nDeduced detail:\n")
        for fo in fo_order:
            cls = classifications[fo]
            if cls['completeness'] == 'deduced':
                f.write(f"{fo}: {cls['completeness_note']}\n")


if __name__ == '__main__':
    main()
